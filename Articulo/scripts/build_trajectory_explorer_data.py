from __future__ import annotations

import json
import os
from pathlib import Path
from typing import Any

os.environ.setdefault("LOKY_MAX_CPU_COUNT", "1")

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn.cluster import KMeans


BASE_DIR = Path(__file__).resolve().parents[2]
DATA_DIR = BASE_DIR / "data"
ARTICLE_DATA_DIR = BASE_DIR / "articulo_neonatologia" / "data"
DASHBOARD_PATH = BASE_DIR / "notebooks" / "metricas" / "pacientes_dashboard.csv"
MASTER_XLSX_PATH = DATA_DIR / "KMC-70k-93-2024-Malnutricion-conVel-DATA-SPSS-20250322.xlsx"
DICTIONARY_XLSX_PATH = DATA_DIR / "KMC-70K-diccionarioVARS-Malnutricion-PhETI-rev20250520-MAIA.xlsx"
OUTPUT_PATH = ARTICLE_DATA_DIR / "trajectory-explorer-data.js"


PHASES = [
    {
        "id": "F0",
        "phase_key": "F0_Prenatal_Parto",
        "label": "Prenatal y parto",
        "time": "Antes del nacimiento",
        "fields": [
            "CP_TallaMadre",
            "CP_TallaPadre",
            "CSP_IngresoMensual",
            "PA_NumDosisCorticoides",
            "BMImadre",
            "CP_PesoPadre",
        ],
    },
    {
        "id": "F1",
        "phase_key": "F1_Nacimiento",
        "label": "Nacimiento",
        "time": "Primeras mediciones al nacer",
        "fields": [
            "zscoretalla0",
            "zscorepeso0",
            "ERN_Peso",
            "CP_TallaMadre",
            "CSP_IngresoMensual",
            "CP_TallaPadre",
        ],
    },
    {
        "id": "F2",
        "phase_key": "F2_Hospitalizacion",
        "label": "Hospitalización",
        "time": "Egreso hospitalario",
        "fields": [
            "zscoretalla0",
            "zscorepeso0",
            "CP_TallaMadre",
            "ERN_Peso",
            "HD_PesoSalida",
            "CP_TallaPadre",
        ],
    },
    {
        "id": "F3",
        "phase_key": "F3_40semanas",
        "label": "40 semanas EC",
        "time": "Entrada al seguimiento corregido",
        "fields": [
            "zscoretalla1",
            "Indexnutricion40sem",
            "BMI1",
            "CP_TallaMadre",
            "gananciatallanacertallaentradaPMC",
            "zscoretalla0",
        ],
    },
    {
        "id": "F4",
        "phase_key": "F4_3meses",
        "label": "3 meses EC",
        "time": "Primer control ambulatorio fuerte",
        "fields": [
            "zscoretalla2",
            "velocidadzscore3m_40semOMS",
            "velocidadzscorepeso40_3m",
            "zscoretalla0",
            "zscoretallaOMS2",
            "CP_TallaMadre",
        ],
    },
    {
        "id": "F5",
        "phase_key": "F5_6meses",
        "label": "6 meses EC",
        "time": "Ventana de aceleración del desempeño",
        "fields": [
            "zscoretalla6",
            "zscorepeso6",
            "zscoretalla6cat",
            "zscorepesotalla6",
            "zscoretalla2",
            "velocidad6_3mesesOMS",
        ],
    },
    {
        "id": "F6",
        "phase_key": "F6_9meses",
        "label": "9 meses EC",
        "time": "Última fase antes del desenlace",
        "fields": [
            "zscoretalla9",
            "zscoretalla6",
            "zscorepeso9",
            "velocidad9_6mesesOMS",
            "zscoretalla9cat",
            "zscorepesotalla9",
        ],
    },
]


FIELD_LABEL_OVERRIDES = {
    "BMImadre": "IMC de la madre",
    "CP_PesoPadre": "Peso del padre",
    "PA_NumDosisCorticoides": "Número de dosis de corticoides en preparto",
    "CSP_IngresoMensual": "Ingreso mensual en el hogar",
    "gestasal": "Edad gestacional al nacer",
    "zscoretalla0": "z-score de talla al nacer",
    "zscorepeso0": "z-score de peso al nacer",
    "HD_PesoSalida": "Peso de salida de hospitalización",
    "HD_TotalDiasHospital": "Días totales de hospitalización",
    "HD_DiasOxigeno": "Días con oxígeno suplementario",
    "gananciatallanacertallaentradaPMC": "Ganancia de talla nacimiento → entrada PMC",
    "BMI1": "IMC a 40 semanas",
    "Indexnutricion40sem": "Crecimiento armónico a 40 semanas",
    "zscoretalla1": "z-score de talla a 40 semanas",
    "zscorepeso1": "z-score de peso a 40 semanas",
    "zscorePC1": "z-score de perímetro cefálico a 40 semanas",
    "RCEUFentonentrada": "RCEU al ingreso al PMC",
    "zscoretalla2": "z-score de talla a 3 meses",
    "zscoretallaOMS2": "z-score de talla OMS a 3 meses",
    "zscorepeso2": "z-score de peso a 3 meses",
    "velocidadzscore3m_40semOMS": "Velocidad de crecimiento 40 semanas → 3 meses",
    "velocidadzscorepeso40_3m": "Velocidad de peso 40 semanas → 3 meses",
    "zscoretalla6": "z-score de talla a 6 meses",
    "zscorepeso6": "z-score de peso a 6 meses",
    "zscoretalla6cat": "Categoría de z-score de talla a 6 meses",
    "zscorepesotalla6": "z-score de peso/talla a 6 meses",
    "velocidad6_3mesesOMS": "Velocidad de crecimiento 3 → 6 meses",
    "zscoretalla9": "z-score de talla a 9 meses",
    "zscorepeso9": "z-score de peso a 9 meses",
    "zscoretalla9cat": "Categoría de z-score de talla a 9 meses",
    "zscorepesotalla9": "z-score de peso/talla a 9 meses",
    "velocidad9_6mesesOMS": "Velocidad de crecimiento 6 → 9 meses",
}


FIELD_IMPORTANCE_BY_PHASE = {
    "F0": {
        "CP_TallaMadre": 11.8,
        "CP_TallaPadre": 9.3,
        "CSP_IngresoMensual": 8.7,
        "PA_NumDosisCorticoides": 7.5,
        "BMImadre": 6.9,
        "CP_PesoPadre": 6.4,
    },
    "F1": {
        "zscoretalla0": 18.7,
        "zscorepeso0": 7.3,
        "ERN_Peso": 7.1,
        "CP_TallaMadre": 5.6,
        "CSP_IngresoMensual": 4.3,
        "CP_TallaPadre": 4.3,
    },
    "F2": {
        "zscoretalla0": 18.0,
        "zscorepeso0": 7.3,
        "CP_TallaMadre": 5.4,
        "ERN_Peso": 4.9,
        "HD_PesoSalida": 4.3,
        "CP_TallaPadre": 3.6,
    },
    "F3": {
        "zscoretalla0": 13.1,
        "zscoretalla1": 4.5,
        "Indexnutricion40sem": 4.3,
        "BMI1": 3.6,
        "CP_TallaMadre": 3.5,
        "gananciatallanacertallaentradaPMC": 3.5,
    },
    "F4": {
        "zscoretalla2": 21.4,
        "velocidadzscore3m_40semOMS": 7.5,
        "velocidadzscorepeso40_3m": 6.3,
        "zscoretalla0": 3.8,
        "zscoretallaOMS2": 3.4,
        "CP_TallaMadre": 2.4,
    },
    "F5": {
        "zscoretalla6": 52.0,
        "zscorepeso6": 7.9,
        "zscoretalla6cat": 3.1,
        "zscorepesotalla6": 2.8,
        "zscoretalla2": 2.5,
        "velocidad6_3mesesOMS": 1.7,
    },
    "F6": {
        "zscoretalla9": 53.8,
        "zscoretalla6": 15.7,
        "zscorepeso9": 2.4,
        "velocidad9_6mesesOMS": 2.1,
        "zscoretalla9cat": 1.6,
        "zscorepesotalla9": 1.0,
    },
}


SUMMARY_FIELDS = [
    "ERN_Peso",
    "ERN_Talla",
    "gestasal",
    "CP_TallaMadre",
    "CP_TallaPadre",
]


BIRTH_GROUP_FIELD = "grupo_nacimiento_cod"


OUTCOMES = [
    {
        "key": "Stunting",
        "label": "Talla baja",
        "shortLabel": "Talla baja",
        "observedColumn": "stunting12m",
        "probabilityPrefix": "prob_Stunting",
        "positiveLabel": "con talla baja",
        "negativeLabel": "sin talla baja",
    },
    {
        "key": "Bajo_peso",
        "label": "Bajo peso",
        "shortLabel": "Bajo peso",
        "observedColumn": "underweight12m_b",
        "probabilityPrefix": "prob_Bajo_peso",
        "positiveLabel": "con bajo peso",
        "negativeLabel": "sin bajo peso",
    },
    {
        "key": "Wasting",
        "label": "Desnutrición aguda",
        "shortLabel": "Aguda",
        "observedColumn": "wasting12m",
        "probabilityPrefix": "prob_Wasting",
        "positiveLabel": "con desnutrición aguda",
        "negativeLabel": "sin desnutrición aguda",
    },
    {
        "key": "Mixta",
        "label": "Condición mixta",
        "shortLabel": "Mixta",
        "observedColumn": "mixta12m",
        "probabilityPrefix": "prob_Mixta",
        "positiveLabel": "con condición mixta",
        "negativeLabel": "sin condición mixta",
    },
]


def probability_columns(outcome: dict[str, str]) -> list[str]:
    return [f"{outcome['probabilityPrefix']}_{phase['phase_key']}" for phase in PHASES]


def load_dictionary() -> dict[str, dict[str, Any]]:
    workbook = load_workbook(DICTIONARY_XLSX_PATH, read_only=True, data_only=True)
    worksheet = workbook["VARS-(KMC70k)"]
    rows = worksheet.iter_rows(min_row=2, values_only=True)
    dictionary: dict[str, dict[str, Any]] = {}
    for row in rows:
        source_name = row[1]
        if not source_name:
            continue
        dictionary[str(source_name)] = {
            "label": FIELD_LABEL_OVERRIDES.get(str(source_name)) or row[3] or source_name,
            "description": row[14] or row[3] or source_name,
            "unit": row[17],
        }
    return dictionary


def load_master_columns(columns: list[str]) -> pd.DataFrame:
    workbook = load_workbook(MASTER_XLSX_PATH, read_only=True, data_only=True)
    worksheet = workbook["Sheet1"]
    header = list(next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    column_positions = {
        str(name): idx
        for idx, name in enumerate(header)
        if name is not None and str(name) in columns
    }
    ordered_columns = [column for column in columns if column in column_positions]
    ordered_indexes = [column_positions[column] for column in ordered_columns]

    rows: list[list[Any]] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        rows.append([row[idx] if idx < len(row) else None for idx in ordered_indexes])

    frame = pd.DataFrame(rows, columns=ordered_columns)
    return frame.replace({"#NULL!": pd.NA})


def clean_value(value: Any) -> Any:
    if pd.isna(value):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float):
            return round(value, 4)
        return value
    return value


def format_field(key: str, row: pd.Series, dictionary: dict[str, dict[str, Any]]) -> dict[str, Any]:
    meta = dictionary.get(key, {})
    return {
        "key": key,
        "label": meta.get("label", FIELD_LABEL_OVERRIDES.get(key, key)),
        "description": meta.get("description", FIELD_LABEL_OVERRIDES.get(key, key)),
        "unit": meta.get("unit"),
        "value": clean_value(row.get(key)),
    }


def field_template(key: str, dictionary: dict[str, dict[str, Any]], phase_id: str | None = None) -> dict[str, Any]:
    meta = dictionary.get(key, {})
    template = {
        "key": key,
        "label": meta.get("label", FIELD_LABEL_OVERRIDES.get(key, key)),
        "description": meta.get("description", FIELD_LABEL_OVERRIDES.get(key, key)),
        "unit": meta.get("unit"),
    }
    if phase_id is not None:
        importance = FIELD_IMPORTANCE_BY_PHASE.get(phase_id, {}).get(key)
        if importance is not None:
            template["importance"] = importance
            template["importanceLabel"] = "Ganancia relativa en la fase"
    return template


def label_cluster(mean_traj: np.ndarray) -> str:
    start = mean_traj[0]
    mid = mean_traj[3]
    final = mean_traj[-1]
    if final >= 0.5 and start >= 0.45:
        return "Riesgo alto persistente"
    if final >= 0.5 and final - start >= 0.2:
        return "Riesgo ascendente"
    if final < 0.25 and start < 0.45:
        return "Riesgo bajo persistente"
    if final < 0.35 and start - final >= 0.2:
        return "Riesgo descendente"
    if mid >= 0.5 and final < 0.5:
        return "Riesgo transitorio"
    return "Riesgo intermedio"


def build_outcome_clusters(frame: pd.DataFrame, outcome: dict[str, str], n_clusters: int = 3) -> dict[int, dict[str, Any]]:
    outcome_column = outcome["observedColumn"]
    probability_cols = probability_columns(outcome)
    valid = frame[[outcome_column, *probability_cols]].copy()
    valid[outcome_column] = pd.to_numeric(valid[outcome_column], errors="coerce")
    for column in probability_cols:
        valid[column] = pd.to_numeric(valid[column], errors="coerce")
    valid = valid.dropna(subset=[outcome_column, *probability_cols])

    if len(valid) < n_clusters:
        return {}

    matrix = valid[probability_cols].to_numpy(dtype=float)
    raw_labels = KMeans(n_clusters=n_clusters, random_state=42, n_init=20).fit_predict(matrix)
    outcome_values = valid[outcome_column].to_numpy(dtype=float)

    ordered_raw_labels = sorted(
        range(n_clusters),
        key=lambda cluster_id: -outcome_values[raw_labels == cluster_id].mean(),
    )
    remap = {raw_label: index + 1 for index, raw_label in enumerate(ordered_raw_labels)}
    mapped_labels = np.array([remap[label] for label in raw_labels])

    profiles: dict[int, dict[str, Any]] = {}
    valid_indexes = list(valid.index)
    for cluster_id in range(1, n_clusters + 1):
        mask = mapped_labels == cluster_id
        mean_traj = matrix[mask].mean(axis=0)
        profile = {
            "cluster": cluster_id,
            "cluster_label": label_cluster(mean_traj),
            "n_pacientes": int(mask.sum()),
            "prevalencia": round(float(outcome_values[mask].mean()), 4),
        }
        for phase, value in zip(PHASES, mean_traj):
            profile[f"prob_media_{phase['phase_key']}"] = round(float(value), 4)
        profiles[cluster_id] = profile

    assignments = {
        int(frame.loc[index, "Idenfinal"]): {
            "cluster": int(mapped_labels[position]),
            "clusterLabel": profiles[int(mapped_labels[position])]["cluster_label"],
        }
        for position, index in enumerate(valid_indexes)
    }
    return {"assignments": assignments, "profiles": list(profiles.values())}


def pick_featured_patients(patients: list[dict[str, Any]], outcome_key: str) -> dict[str, int | None]:
    outcome_patients = [
        patient
        for patient in patients
        if patient["outcomes"].get(outcome_key)
    ]
    positives = [patient for patient in outcome_patients if patient["outcomes"][outcome_key]["outcomeReal"] == 1]
    negatives = [patient for patient in outcome_patients if patient["outcomes"][outcome_key]["outcomeReal"] == 0]

    if not positives or not negatives:
        first_id = outcome_patients[0]["id"] if outcome_patients else None
        return {
            "earlyHigh": first_id,
            "stableLow": first_id,
            "lateEscalation": first_id,
        }

    early_high = max(
        positives,
        key=lambda patient: (
            patient["outcomes"][outcome_key]["probs"][0] >= 0.5,
            patient["outcomes"][outcome_key]["probs"][-1],
            patient["outcomes"][outcome_key]["probs"][0],
        ),
    )
    stable_low = min(
        negatives,
        key=lambda patient: (
            patient["outcomes"][outcome_key]["probs"][-1],
            patient["outcomes"][outcome_key]["probs"][0],
        ),
    )
    late_escalation = max(
        positives,
        key=lambda patient: (
            patient["outcomes"][outcome_key]["probs"][5] - patient["outcomes"][outcome_key]["probs"][3],
            patient["outcomes"][outcome_key]["probs"][-1] - patient["outcomes"][outcome_key]["probs"][0],
        ),
    )
    return {
        "earlyHigh": early_high["id"],
        "stableLow": stable_low["id"],
        "lateEscalation": late_escalation["id"],
    }


def build_dataset() -> dict[str, Any]:
    dictionary = load_dictionary()
    summary_templates = [field_template(field, dictionary) for field in SUMMARY_FIELDS]
    phase_templates = {
        phase["id"]: [field_template(field, dictionary, phase["id"]) for field in phase["fields"]]
        for phase in PHASES
    }

    dashboard_columns = [
        "Idenfinal",
        "Iden_Sede",
        "ano",
        "era_covid",
        *(outcome["observedColumn"] for outcome in OUTCOMES),
        BIRTH_GROUP_FIELD,
        *(column for outcome in OUTCOMES for column in probability_columns(outcome)),
        *SUMMARY_FIELDS,
        *(field for phase in PHASES for field in phase["fields"]),
    ]
    merged = pd.read_csv(
        DASHBOARD_PATH,
        usecols=lambda column: column in set(dashboard_columns),
        low_memory=False,
    )
    merged = merged.replace({"#NULL!": pd.NA})
    for outcome in OUTCOMES:
        merged[outcome["observedColumn"]] = pd.to_numeric(merged[outcome["observedColumn"]], errors="coerce")

    cluster_payloads = {
        outcome["key"]: build_outcome_clusters(merged, outcome)
        for outcome in OUTCOMES
    }

    patients: list[dict[str, Any]] = []
    for patient_idx, row in merged.iterrows():
        outcomes_payload: dict[str, Any] = {}
        patient_id = int(row["Idenfinal"])
        for outcome in OUTCOMES:
            outcome_key = outcome["key"]
            observed_value = row.get(outcome["observedColumn"])
            probs = [clean_value(row.get(column)) for column in probability_columns(outcome)]
            if pd.isna(observed_value) or any(prob is None for prob in probs):
                continue
            cluster_assignment = cluster_payloads[outcome_key]["assignments"].get(patient_id, {})
            outcomes_payload[outcome_key] = {
                "outcomeReal": int(observed_value),
                "finalRisk": probs[-1],
                "probs": probs,
                "cluster": clean_value(cluster_assignment.get("cluster")),
                "clusterLabel": clean_value(cluster_assignment.get("clusterLabel")),
            }

        if not outcomes_payload:
            continue

        patient = {
            "patientIdx": int(patient_idx),
            "id": patient_id,
            "sede": clean_value(row.get("Iden_Sede")),
            "periodo": clean_value(row.get("ano")),
            "eraCovid": clean_value(row.get("era_covid")),
            "birthGroup": clean_value(row.get(BIRTH_GROUP_FIELD)),
            "summaryValues": [clean_value(row.get(field)) for field in SUMMARY_FIELDS],
            "phaseValues": [
                [clean_value(row.get(field)) for field in phase["fields"]]
                for phase in PHASES
            ],
            "outcomes": outcomes_payload,
        }
        patients.append(patient)

    outcome_meta = {}
    featured = {}
    for outcome in OUTCOMES:
        outcome_key = outcome["key"]
        outcome_patients = [patient for patient in patients if patient["outcomes"].get(outcome_key)]
        positives = sum(1 for patient in outcome_patients if patient["outcomes"][outcome_key]["outcomeReal"] == 1)
        outcome_meta[outcome_key] = {
            **outcome,
            "totalPatients": len(outcome_patients),
            "positivePatients": positives,
            "negativePatients": len(outcome_patients) - positives,
            "clusterProfiles": cluster_payloads[outcome_key]["profiles"],
        }
        featured[outcome_key] = pick_featured_patients(patients, outcome_key)

    return {
        "meta": {
            "title": "Explorador interactivo de trayectorias nutricionales",
            "source": "notebooks/metricas/pacientes_dashboard.csv",
            "datasetLabel": "Todos los pacientes con desenlaces a 12 meses, predicciones por fase y clusters por desenlace",
            "defaultOutcome": "Stunting",
            "totalPatients": len(patients),
            "outcomes": outcome_meta,
            "phases": [
                {
                    "id": phase["id"],
                    "label": phase["label"],
                    "time": phase["time"],
                    "phaseKey": phase["phase_key"],
                }
                for phase in PHASES
            ],
            "summaryFields": summary_templates,
            "phaseFieldTemplates": phase_templates,
            "birthGroupField": field_template(BIRTH_GROUP_FIELD, dictionary),
        },
        "featured": featured,
        "patients": patients,
    }


def main() -> None:
    dataset = build_dataset()
    ARTICLE_DATA_DIR.mkdir(parents=True, exist_ok=True)
    payload = "window.TRAJECTORY_EXPLORER_DATA = " + json.dumps(dataset, ensure_ascii=False) + ";\n"
    temp_path = OUTPUT_PATH.with_suffix(".tmp")
    temp_path.write_text(payload, encoding="utf-8")
    temp_path.replace(OUTPUT_PATH)
    print(f"Archivo generado: {OUTPUT_PATH}")
    print(
        f"Pacientes exportados: {dataset['meta']['totalPatients']} "
        f"(universo con al menos un desenlace completo)"
    )


if __name__ == "__main__":
    main()
